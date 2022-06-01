import xlwt
import numpy as np
import openpyxl
from openpyxl import Workbook

header = ['Spalte 0', 'Spalte 1', 'Spalte 2']

x = np.zeros((27,3))

book = xlwt.Workbook(encoding='utf-8', style_compression=0)
sheet = book.add_sheet('test', cell_overwrite_ok=True)

i=0
for k in header:
    sheet.write(0,i,k) # Zeile, Spalte, Inhalt
    i=i+1

zeile = 1
spalte = 0

while zeile <= 27:
    while spalte < 3:
        sheet.write(zeile,spalte, x[zeile-1][spalte])
        spalte = spalte + 1
    zeile = zeile + 1
    spalte = 0

book.save('datasets/test_toexcel.xls')

#workbook = Workbook()
#booksheet = workbook.active
workbook = openpyxl.Workbook()
# Aufrufen sheet1 = workbook['excited']
sheet0 = workbook.active
sheet1 = workbook.create_sheet('AD')

sheet1.cell(1,1).value = 6 ###############
i=1
while i < 3000:
    sheet0.cell(1,i).value = i
    sheet1.cell(i,1).value = i
    i=i+1

#booksheet.append([11,87])

workbook.save('datasets/test_openpyxl.xlsx')