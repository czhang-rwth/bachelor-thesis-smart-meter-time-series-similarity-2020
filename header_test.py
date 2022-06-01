import pandas as pd # This is always assumed but is included here as an introduction.
import numpy as np
import matplotlib.pyplot as plt
from scipy import optimize
import openpyxl
import xlwt

header = ['']*7 #2701 + 1
k = 0
i = 1
ev = 0
while ev < 2:#2701 hier egal
    if ev == 0:
        k = 0
        header[k] = 'Aufloesung in Minuten'
    else:
        while i <= 4:  # 74
            j = i + 1
            while j <= 4:  # 74
                header[k] = 'Haushalt ' + str(i) + '&' + str(j)
                j = j + 1
                k=k+1
            i = i + 1
    ev = ev + 1

workbook = openpyxl.Workbook()
sheet_corr = workbook.create_sheet('Correlation')
sheet_mean = workbook.create_sheet('Mean_of_distance')
sheet_std = workbook.create_sheet('Standardabweichung_of_distance')
sheet_corr_log = workbook.create_sheet('log_Anpassung_corr')
sheet_mean_log = workbook.create_sheet('log_Anpassung_mean')
sheet_std_log = workbook.create_sheet('log_Anpassung_std')

zeile = 1
spalte = 1
while zeile <= 121:
    while spalte <= 7:
        if zeile==1:
            sheet_corr.cell(zeile, spalte).value = header[spalte-1]
            sheet_mean.cell(zeile, spalte).value = header[spalte-1]
            sheet_std.cell(zeile, spalte).value = header[spalte-1]
        else:
            sheet_corr.cell(zeile, spalte).value = 1
            sheet_mean.cell(zeile, spalte).value = 1
            sheet_std.cell(zeile, spalte).value = 1
        spalte = spalte + 1
    zeile = zeile + 1
    spalte = 1

workbook.save('datasets/header_test.xlsx')