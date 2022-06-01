import pandas as pd # This is always assumed but is included here as an introduction.
import numpy as np
import matplotlib.pyplot as plt
from scipy import optimize
import openpyxl
import xlwt
from scipy import signal

def indexieren_htw_xlsx(csv_Datei ,excel_Datei):
    haushalt_csv = pd.read_csv(csv_Datei)
    ### Zwischenschritt print(haushalt_csv)
    haushalt_csv = haushalt_csv['Timestamp;PL1[W];PL2[W];PL3[W];QL1[W];QL2[W];QL3[W]'].str.split(';', expand=True)
    ### Zwischenschritt print(haushalt_csv)
    haushalt_csv.to_excel(excel_Datei)
    haushalt = pd.read_excel(excel_Datei)
    len_0 = len(haushalt)
    ### Zwischenschritt print(haushalt)

    haushalt.rename(columns={'Unnamed: 0': 'Tage'}, inplace=True)
    haushalt.rename(columns={0: 'Zeit'}, inplace=True)
    haushalt.rename(columns={1: 'PL1[W]'}, inplace=True)
    haushalt.rename(columns={2: 'PL2[W]'}, inplace=True)
    haushalt.rename(columns={3: 'PL3[W]'}, inplace=True)
    haushalt.rename(columns={4: 'QL1[W]'}, inplace=True)
    haushalt.rename(columns={5: 'QL2[W]'}, inplace=True)
    haushalt.rename(columns={6: 'QL3[W]'}, inplace=True)

    #haushalt.index = pd.to_datetime(haushalt['Tage'] + haushalt['Zeit']) ###Fehler: index die ganze Zeitreihe wirkt danach nicht mehr
    haushalt['Zeitreihe'] = pd.Series((haushalt['Tage'] + haushalt['Zeit']), index=haushalt.index)
    haushalt['Wirkleistung PL[kW]'] = pd.Series((haushalt['PL1[W]'] + haushalt['PL2[W]'] + haushalt['PL3[W]']) * 0.001,
                                                index=haushalt.index)
    haushalt = haushalt.drop(columns=['Tage', 'Zeit','PL1[W]', 'PL2[W]', 'PL3[W]', 'QL1[W]', 'QL2[W]', 'QL3[W]'])
    ### Zwischenschritt print(haushalt)
    #print(haushalt['Wirkleistung PL[kW]'].dtypes)
    #haushalt[column] = haushalt[column].astype(int)
    index_neu = pd.date_range('01.01.2010', '01.01.2011', freq='1min', closed='left')
    haushalt_neu = pd.Series(np.arange(len_0), index=index_neu, dtype=float)
    #####haushalt_neu = pd.Series(haushalt['Wirkleistung PL[kW]'], index=index_neu, dtype=float) ###Fehler: haushalt_neu zeigt dann immer NaN
    i = 0
    while i < len_0:
        haushalt_neu[i] = haushalt['Wirkleistung PL[kW]'][i]
        ###print(i, " ", haushalt['Wirkleistung PL[kW]'][i]," ", haushalt_neu[i] )
        i=i+1

    print(haushalt_neu)
    return haushalt_neu

def plotten_glatt_blau(x, y, title='', xlabel='', ylabel='', label='', dpi=100):
    plt.figure(figsize=(16, 5))
    plt.xticks(fontsize=20)
    plt.yticks(fontsize=20)

    # plt.plot(x, y, 's-',label=label)
    plt.plot(x, y, label=label)
    plt.xlabel(xlabel, fontsize=20)
    plt.ylabel(ylabel, fontsize=20)
    plt.title(title, fontsize=20, ha='center')
    plt.legend(loc='best', fontsize=17)  # fuer label
    plt.show()

def plotten_glatt_rot(x, y, title='', xlabel='', ylabel='', label='', dpi=100):
    plt.figure(figsize=(16, 5))
    plt.xticks(fontsize=20)
    plt.yticks(fontsize=20)

    #plt.plot(x, y, 's-',label=label)
    plt.plot(x, y, color='red', label=label)
    plt.xlabel(xlabel, fontsize=20)
    plt.ylabel(ylabel, fontsize=20)
    plt.title(title, fontsize=20, ha='center')
    plt.legend(loc='best', fontsize=17)
    plt.show()

def plot_zwei_kurven_glatt(x1, y1, x2, y2, title, xlabel, ylabel, label1, label2, png_datei, dpi=100):
    plt.figure(figsize=(16, 6))
    plt.xticks(fontsize=20)
    plt.yticks(fontsize=20)

    plt.plot(x1, y1, color='blue', label=label1)
    plt.plot(x2, y2, color='red', label=label2)
    #plt.plot(haushalt1.index, haushalt1, 'o-', color='red', label=label1)
    #plt.plot(haushalt2.index, haushalt2, 's--', color='blue', label=label2)
    plt.xlabel(xlabel, fontsize=20)
    plt.ylabel(ylabel, fontsize=20)
    plt.title(title, fontsize=20, ha='center')
    plt.legend(loc='best', fontsize=17)  # fuer label
    plt.savefig(png_datei, bbox_inches='tight')
    #plt.savefig('datasets/test.png', bbox_inches='tight')
    plt.show()

def haushalt_einlesen():
    i=1
    haushalt=[0]*74
    while i<=74:
        print('Einlesen Haushalt'+str(i))
        haushalt[i-1] = indexieren_htw_xlsx('datasets/HTW Berlin/1min_htwBerlin/Haushalt_' +str(i)+'.csv' ,
                                'datasets/HTW Berlin/1min_htwBerlin/Haushalt_'+str(i)+'.xlsx')
        i=i+1
    return haushalt

def haushalt_einlesen_test():
    i=1
    haushalt=[0]*4
    while i<=4:
        print('Einlesen Haushalt'+str(i))
        haushalt[i-1] = indexieren_htw_xlsx('datasets/HTW Berlin/1min_htwBerlin/Haushalt_' +str(i)+'.csv' ,
                                'datasets/HTW Berlin/1min_htwBerlin/Haushalt_'+str(i)+'.xlsx')
        i=i+1
    return haushalt

def xcorr_adavanced():
    haushalt = haushalt_einlesen()
    global xcorr
    xcorr_index = 1
    i = 1

    while i <= 74:
        j = i + 1
        while j <= 74:
            xcorr[xcorr_index] = signal.correlate(haushalt[i - 1], haushalt[j - 1], mode='full', method='direct')
            print('Haushalt'+str(i)+' und Haushalt'+str(j))
            print(xcorr[xcorr_index])
            xcorr_index = xcorr_index + 1
            j = j + 1
        i = i + 1

xcorr = [0] * 2702 #2701+1

# Excel design
header = ['']*2702 #2701 + 1
ev=0
k = 0
i = 1
while ev < 2:#2701 hier egal
    if ev == 0:
        k=0
        header[k] = 'Verschiebung in Minuten'
        k=k+1
    else:
        while i <= 74:  # 74
            j = i + 1
            while j <= 74:  # 74
                header[k] = 'Haushalt ' + str(i) + '&' + str(j)
                j = j + 1
                k=k+1
            i = i + 1
    ev=ev+1

xcorr[0]= np.arange(1,1051199+1,step=1)

xcorr_adavanced()

# export in excel
workbook = openpyxl.Workbook()
sheet_xcorr = workbook.create_sheet('CrossCorrelation')

zeile = 1
spalte = 1
while spalte <= 2702:  # 2701+1
    sheet_xcorr.cell(zeile, spalte).value = header[spalte - 1]
    spalte = spalte + 1
spalte=1

while spalte <= 2702: # 2701 +1
    while zeile <= 1000:
        sheet_xcorr.cell(zeile+1, spalte).value = xcorr[spalte-1][zeile-1]
        zeile = zeile + 1
    spalte=spalte+1
    zeile = 1

workbook.save('datasets/Ergebnisse_xcorr_alle.xlsx')