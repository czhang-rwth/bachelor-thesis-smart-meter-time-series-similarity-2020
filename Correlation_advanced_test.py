import pandas as pd # This is always assumed but is included here as an introduction.
import numpy as np
import matplotlib.pyplot as plt
from scipy import optimize
import openpyxl
import xlwt

def indexieren_htw_xlsx(csv_Datei ,excel_Datei):
    haushalt_csv = pd.read_csv(csv_Datei)
    ### Zwischenschritt print(haushalt_csv)
    haushalt_csv = haushalt_csv['Timestamp;PL1[W];PL2[W];PL3[W];QL1[W];QL2[W];QL3[W]'].str.split(';', expand=True)
    ### Zwischenschritt print(haushalt_csv)
    #haushalt_csv.to_excel(excel_Datei)
    haushalt = pd.read_excel(excel_Datei)
    len_0 = len(haushalt)
    ### Zwischenschritt print(haushalt)

    haushalt.rename(columns={'Unnamed: 0': 'Tage'}, inplace=True)
    haushalt.rename(columns={0: 'Zeit'}, inplace=True)
    haushalt.rename(columns={1: 'PL1[W]'}, inplace=True)
    haushalt.rename(columns={2: 'PL2[W]'}, inplace=True)
    haushalt.rename(columns={3: 'PL3[W]'}, inplace=True)
    haushalt.rename(columns={4: 'QL1[W]'}, inplace=True)
    haushalt.rename(columns={5: 'QL2[W]'}, inplace=True)
    haushalt.rename(columns={6: 'QL3[W]'}, inplace=True)

    #haushalt.index = pd.to_datetime(haushalt['Tage'] + haushalt['Zeit']) ###Fehler: index die ganze Zeitreihe wirkt danach nicht mehr
    haushalt['Zeitreihe'] = pd.Series((haushalt['Tage'] + haushalt['Zeit']), index=haushalt.index)
    haushalt['Wirkleistung PL[kW]'] = pd.Series((haushalt['PL1[W]'] + haushalt['PL2[W]'] + haushalt['PL3[W]']) * 0.001,
                                                index=haushalt.index)
    haushalt = haushalt.drop(columns=['Tage', 'Zeit','PL1[W]', 'PL2[W]', 'PL3[W]', 'QL1[W]', 'QL2[W]', 'QL3[W]'])
    ### Zwischenschritt print(haushalt)
    #print(haushalt['Wirkleistung PL[kW]'].dtypes)
    #haushalt[column] = haushalt[column].astype(int)
    index_neu = pd.date_range('01.01.2010', '01.01.2011', freq='1min', closed='left')
    haushalt_neu = pd.Series(np.arange(len_0), index=index_neu, dtype=float)
    #####haushalt_neu = pd.Series(haushalt['Wirkleistung PL[kW]'], index=index_neu, dtype=float) ###Fehler: haushalt_neu zeigt dann immer NaN
    i = 0
    while i < len_0:
        haushalt_neu[i] = haushalt['Wirkleistung PL[kW]'][i]
        ###print(i, " ", haushalt['Wirkleistung PL[kW]'][i]," ", haushalt_neu[i] )
        i=i+1

    print(haushalt_neu)
    return haushalt_neu

def plotten_glatt_blau(x, y, title='', xlabel='', ylabel='', label='', dpi=100):
    plt.figure(figsize=(16, 5))
    plt.xticks(fontsize=20)
    plt.yticks(fontsize=20)

    # plt.plot(x, y, 's-',label=label)
    plt.plot(x, y, label=label)
    plt.xlabel(xlabel, fontsize=20)
    plt.ylabel(ylabel, fontsize=20)
    plt.title(title, fontsize=20, ha='center')
    plt.legend(loc='best', fontsize=17)  # fuer label
    plt.show()

def plotten_glatt_rot(x, y, title='', xlabel='', ylabel='', label='', dpi=100):
    plt.figure(figsize=(16, 5))
    plt.xticks(fontsize=20)
    plt.yticks(fontsize=20)

    #plt.plot(x, y, 's-',label=label)
    plt.plot(x, y, color='red', label=label)
    plt.xlabel(xlabel, fontsize=20)
    plt.ylabel(ylabel, fontsize=20)
    plt.title(title, fontsize=20, ha='center')
    plt.legend(loc='best', fontsize=17)
    plt.show()

def plot_zwei_kurven_glatt(x1, y1, x2, y2, title, xlabel, ylabel, label1, label2, png_datei, dpi=100):
    plt.figure(figsize=(16, 6))
    plt.xticks(fontsize=20)
    plt.yticks(fontsize=20)

    plt.plot(x1, y1, color='blue', label=label1)
    plt.plot(x2, y2, color='red', label=label2)
    #plt.plot(haushalt1.index, haushalt1, 'o-', color='red', label=label1)
    #plt.plot(haushalt2.index, haushalt2, 's--', color='blue', label=label2)
    plt.xlabel(xlabel, fontsize=20)
    plt.ylabel(ylabel, fontsize=20)
    plt.title(title, fontsize=20, ha='center')
    plt.legend(loc='best', fontsize=17)  # fuer label
    plt.savefig(png_datei, bbox_inches='tight')
    #plt.savefig('datasets/test.png', bbox_inches='tight')
    plt.show()


def korrelation(haushalt1, haushalt2, m, n, c):
    i=1
    t='T'

    #Minuten 1 min bis 3 Tage
    werte_corr = [0]*120
    werte_mean = [0]*120
    werte_std = [0]*120

    global corr_gesamt
    global mean_gesamt
    global std_gesamt

    while i <= 120:
        j = str(i)
        distanz = manhattan_distance(haushalt1.resample(j+t).mean(), haushalt2.resample(j+t).mean())
        werte_corr[i-1] = haushalt1.resample(j+t).mean().corr(haushalt2.resample(j+t).mean(), method='pearson')
        corr_gesamt[i-1][c] = werte_corr[i-1]

        werte_mean[i-1] = np.mean(distanz)
        mean_gesamt[i-1][c] = werte_mean[i-1]

        werte_std[i-1] = np.std(distanz)
        std_gesamt[i-1][c] = werte_std[i-1]

        print('i:', i, ' corr: ',werte_corr[i-1])
        i=i+1

    achse_x_min = np.arange(1, 121, step=1)
    achse_x_func = np.arange(1,121)

    A_log_corr, B_log_corr, C_log_corr = optimize.curve_fit(f_log, achse_x_min, werte_corr)[0]
    log_func_corr = A_log_corr * np.log(achse_x_func + B_log_corr) + C_log_corr
    #print('Koeffizient Logfunktion Haushalt_Corr', A_log_corr, B_log_corr, C_log_corr)

    A_log_mean, B_log_mean, C_log_mean = optimize.curve_fit(f_log, achse_x_min, werte_mean)[0]
    log_func_mean = A_log_mean * np.log(achse_x_func + B_log_mean) + C_log_mean
    #print('Koeffizient Logfunktion Haushalt_mean', A_log_mean, B_log_mean, C_log_mean)

    A_log_std, B_log_std, C_log_std = optimize.curve_fit(f_log, achse_x_min, werte_std)[0]
    log_func_std = A_log_std * np.log(achse_x_func + B_log_std) + C_log_std
    #print('Koeffizient Logfunktion Haushalt_std', A_log_std, B_log_std, C_log_std)

    global koeffizient_corr
    global koeffizient_mean
    global koeffizient_std

    koeffizient_corr[c-1][0] = float(m)
    koeffizient_corr[c - 1][1] = float(n)
    koeffizient_corr[c - 1][2] = A_log_corr
    koeffizient_corr[c - 1][3] = B_log_corr
    koeffizient_corr[c - 1][4] = C_log_corr

    koeffizient_mean[c-1][0] = float(m)
    koeffizient_mean[c - 1][1] = float(n)
    koeffizient_mean[c - 1][2] = A_log_mean
    koeffizient_mean[c - 1][3] = B_log_mean
    koeffizient_mean[c - 1][4] = C_log_mean

    koeffizient_std[c-1][0] = float(m)
    koeffizient_std[c - 1][1] = float(n)
    koeffizient_std[c - 1][2] = A_log_std
    koeffizient_std[c - 1][3] = B_log_std
    koeffizient_std[c - 1][4] = C_log_std

    #plot_zwei_kurven_glatt(achse_x_min, werte_corr, achse_x_func, log_func_corr, title='Korrelation Wirkleistungen Haushalt '+m+'&'+n, xlabel='Zeitaufloesung Minuten',
    #                       ylabel='Korrelationswert', label1='Korrelationskurve', label2='Logarithmische Anpassung Korrelationskurve',
    #                       png_datei='datasets/Bilder/Corr/corrHaushalt'+m+'_'+n+'.png')

    #plot_zwei_kurven_glatt(achse_x_min, werte_mean, achse_x_func, log_func_mean, title='Mittelwert von Distanzen Wirkleistungen Haushalt'+m+'&'+n+' je nach Resamplings', xlabel='Zeitaufloesung Minuten',
    #                       ylabel='Mittelwert', label1='Mittelwertkurve', label2='Logarithmische Anpassung der Mittelwertkurve',
    #                       png_datei='datasets/Bilder/Mean/meanHaushalt'+m+'_'+n+'.png')

    #plot_zwei_kurven_glatt(achse_x_min, werte_std, achse_x_func, log_func_std, title='Standardabweichung von Distanzen Wirkleistungen Haushalt'+m+'&'+n+' je nach Resamplings', xlabel='Zeitaufloesung in Minuten',
    #                       ylabel='Standardabweichung', label1='Standardabweichungskurve', label2='Logarithmische Anpassung der Standardabweichungskurve',
     #                      png_datei='datasets/Bilder/Std/stdHaushalt'+m+'_'+n+'.png')

def f_log(x,A,B,C):
    return A*np.log(x+B)+C

def manhattan_distance(x, y): #, title=''):  #x und y müssen gleiche zeitreihe haben
    i=0
    manhattan=[0]*len(x)

    while i<len(x):
        manhattan[i] = abs(x[i]-y[i])
        i=i+1

    return manhattan

def haushalt_einlesen():
    i=1
    haushalt=[0]*4
    while i<=4:
        print('Einlesen Haushalt'+str(i))
        haushalt[i-1] = indexieren_htw_xlsx('datasets/HTW Berlin/1min_htwBerlin/Haushalt_' +str(i)+'.csv' ,
                                'datasets/HTW Berlin/1min_htwBerlin/Haushalt_'+str(i)+'.xlsx')
        i=i+1
    return haushalt

def korrelation_advanced(haushalt):
    i=1
    counter = 0

    while i<=4:
        j=i+1
        while j<=4:
            counter = counter + 1
            korrelation(haushalt[i-1],haushalt[j-1],str(i),str(j),counter)
            j=j+1
        i=i+1

haushalt = haushalt_einlesen()

# Excel design
header = ['']*7 #2701 + 1
ev=0
k = 0
i = 1
while ev < 2:#2701 hier egal
    if ev == 0:
        k=0
        header[k] = 'Aufloesung in Minuten'
        k=k+1
    else:
        while i <= 4:  # 74
            j = i + 1
            while j <= 4:  # 74
                header[k] = 'Haushalt ' + str(i) + '&' + str(j)
                j = j + 1
                k=k+1
            i = i + 1
    ev=ev+1

corr_gesamt = np.zeros((120,7))
mean_gesamt = np.zeros((120,7))
std_gesamt = np.zeros((120,7))
k = 0
while k < 120:
    corr_gesamt[k][0] = k+1
    mean_gesamt[k][0] = k + 1
    std_gesamt[k][0] = k + 1
    k = k + 1

header_log=['Erster_Haushalt','Zweiter_Haushalt','A','B','C']
koeffizient_corr = np.zeros((6,5))
koeffizient_mean = np.zeros((6, 5))
koeffizient_std = np.zeros((6, 5))

korrelation_advanced(haushalt)

# export in excel
workbook = openpyxl.Workbook()
sheet_corr = workbook.create_sheet('Correlation')
sheet_mean = workbook.create_sheet('Mean_of_distance')
sheet_std = workbook.create_sheet('Standardabweichung_of_distance')
sheet_corr_log = workbook.create_sheet('log_Anpassung_corr')
sheet_mean_log = workbook.create_sheet('log_Anpassung_mean')
sheet_std_log = workbook.create_sheet('log_Anpassung_std')

zeile = 1
spalte = 1
while zeile <= 121:
    while spalte <= 7:
        if zeile==1:
            sheet_corr.cell(zeile, spalte).value = header[spalte-1]
            sheet_mean.cell(zeile, spalte).value = header[spalte-1]
            sheet_std.cell(zeile, spalte).value = header[spalte-1]
        else:
            sheet_corr.cell(zeile, spalte).value = corr_gesamt[zeile-2][spalte-1]
            sheet_mean.cell(zeile, spalte).value = mean_gesamt[zeile -2][spalte-1]
            sheet_std.cell(zeile, spalte).value = std_gesamt[zeile - 2][spalte-1]
        spalte = spalte + 1
    zeile = zeile + 1
    spalte = 1

zeile = 1
spalte = 1
while zeile <=7:
    while spalte <= 5:
        if zeile ==1:
            sheet_corr_log.cell(zeile, spalte).value = header_log[spalte-1]
            sheet_mean_log.cell(zeile, spalte).value = header_log[spalte-1]
            sheet_std_log.cell(zeile, spalte).value = header_log[spalte-1]
        else:
            sheet_corr_log.cell(zeile, spalte).value = koeffizient_corr[zeile-2][spalte-1]
            sheet_mean_log.cell(zeile, spalte).value = koeffizient_mean[zeile-2][spalte-1]
            sheet_std_log.cell(zeile, spalte).value = koeffizient_std[zeile-2][spalte-1]
        spalte=spalte+1
    zeile=zeile+1
    spalte=1

workbook.save('datasets/Ergebnisse_test2.xlsx')